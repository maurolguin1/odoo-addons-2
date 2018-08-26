# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
from openerp import api, models, fields, _
from openerp.exceptions import UserError
import base64
import tempfile
import openpyxl
import datetime


class ImportWorksheet(models.TransientModel):
    _name = "product_upload.import_worksheet"

    data = fields.Binary(
        'File',
        required=True
    )
    name = fields.Char(
        'Filename',
        readonly=True
    )

    def read_data(self, sheet):
        """ Read the spreadsheet into a data structure
        """

        def check(name, col, data_type):
            # chequear que no estemos fuera de rango
            if col > sheet.max_column:
                return {}

            err = _('row %d of sheet %s')

            value = row[col].value
            if data_type == 'str':
                try:
                    value = str(value)
                except Exception as ex:
                    text = _('Not a string') + ' ' + err % row[col].row, sheet.title
                    self.add_error(text)
                finally:
                    return {name: value}
            if data_type == 'float':
                try:
                    value = float(value)
                except Exception as ex:
                    self.add_error('')
                finally:
                    return {name: value}
            if data_type == 'integer':
                try:
                    value = int(value)
                except Exception as ex:
                    self.add_error('')
                finally:
                    return {name: value}

            if data_type == 'boolean':
                value = True if value else False
                return {name: value}

        if sheet.max_column < 4:
            self.add_error(_('Too few columns in sheet %s.') % sheet.title)
            return

        ret = []
        row_number = 0
        for row in sheet.iter_rows(min_row=2, min_col=1,
                                   max_row=sheet.max_row,
                                   max_col=sheet.max_column):
            row_number += 1
            line = {}
            line.update(check('default_code', 0, 'str'))
            line.update(check('cost', 1, 'float'))
            line.update(check('currency', 2, 'str'))
            line.update(check('margin', 3, 'float'))
            line.update(check('name', 4, 'str'))
            line.update(check('purchase_tax', 5, 'float'))
            line.update(check('sale_tax', 6, 'float'))
            line.update(check('barcode', 7, 'integer'))
            line.update(check('meli', 8, 'str'))
            line.update(check('mercadoshops', 9, 'boolean'))

            ret.append(line)
        return ret

    def check_data(self, data, vendor):
        """ Check data structure for errors
        """
        product_obj = self.env['product.product']
        for row in data:
            # check product exists
            domain = [('default_code', '=', row['default_code'])]
            if not product_obj.search(domain):
                raise UserError(
                    _(u'ERROR in line %s, from vendor %s, product '
                      u'"%s" not found') %
                    (row['row'], vendor, row['default_code']))
            try:
                # check list price is a number
                float(row['list_price'])
            except (ValueError, TypeError):
                raise UserError(
                    _(u'Error in line %s, from vendor %s list price "%s" '
                      u'is not a number') %
                    (row['row'], vendor, row['list_price']))
            try:
                # check standard price is a number
                float(row['standard_price'])
            except (ValueError, TypeError):
                raise UserError(
                    _(u'Error in line %s, from vendor %s, standard price '
                      u'"%s" is not a number') %
                    (row['row'], vendor, row['standard_price']))

    def process_data(self, data, vendor):
        """ Process data structure setting prices in system
        """
        product_obj = self.env['product.product']
        for row in data:
            domain = [('default_code', '=', row['default_code'])]
            prod = product_obj.search(domain)
            prod.list_price = float(row['list_price'])
            prod.product_tmpl_id.set_cost(vendor,
                                          float(row['standard_price']),
                                          str(datetime.datetime.now())[0:10])

    @api.multi
    def import_file(self):
        """ Se ejecuta desde el boton import file del wizard, sube la planilla
            y la deja en un archivo temporario
        """
        for rec in self:
            # get log record
            log_obj = self.env['product_upload.log']
            current_id = self.env.context.get('default_active_id', False)
            self.log = log_obj.search([('id', '=', current_id)])
            if not self.log:
                # TODO esto no anda, ver porque
                raise Exception('unknown error')

            data = base64.decodestring(rec.data)
            (fileno, fp_name) = tempfile.mkstemp('.xlsx', 'openerp_')

            # escribir la planilla en un temporario
            with open(fp_name, 'w') as ws:
                ws.write(data)

            self.process_tmp_file(fp_name)

    def process_tmp_file(self, tmp_file):
        """ Procesa la planilla que esta en el archivo temporario
        """

        # leer la planilla del temporario en solo lectura
        wb = openpyxl.load_workbook(filename=tmp_file, read_only=True,
                                    data_only=True)

        # cada hoja de la planilla es un vendor
        partner_obj = self.env['res.partner']
        vendors = wb.sheetnames
        for vendor in vendors:
            partner = partner_obj.search([('ref', '=', vendor)])
            if not partner:
                self.add_error(_('Vendor ref %s not found.') % vendor)
                break

            data = self.read_data(wb[vendor])
            # self.check_data(data, vendor)
            # self.process_data(data, vendor)
            self.log.state = 'done'

    def add_error(self, text):
        """ add an error to log
        """
        self.log.state = 'error'
        self.log.errors += 1
        self.log.error_ids.create({
            'name': text,
            'log_id': self.log.id
        })
